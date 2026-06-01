import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from pathlib import Path
import requests
import json
import re
import base64
import io
from datetime import date
from bs4 import BeautifulSoup
from openpyxl import load_workbook

GITHUB_REPO = "gijsbertvandeweg-stack/qp-price-monitor"
GITHUB_FILE = "data.xlsx"

RETAILER_MAP = {
    "ah.nl": "Albert Heijn",
    "jumbo.com": "Jumbo",
    "aldi.nl": "Aldi",
    "plus.nl": "Plus",
    "dirk.nl": "Dirk",
    "dekamarkt.nl": "DekaMarkt",
    "vomar.nl": "Vomar",
    "hoogvliet.com": "Hoogvliet",
    "poiesz": "Poiesz",
    "spar.nl": "Spar",
}

def detect_retailer(url: str) -> str:
    for domain, name in RETAILER_MAP.items():
        if domain in url:
            return name
    return "Onbekend"

def fetch_price_requests(url: str):
    try:
        session = requests.Session()
        session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36",
            "Accept-Language": "nl-NL,nl;q=0.9",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Sec-Fetch-Mode": "navigate",
        })
        if "ah.nl" in url:
            session.get("https://www.ah.nl/", timeout=10)
        resp = session.get(url, timeout=15, allow_redirects=True)
        if resp.status_code != 200:
            return None, f"Mislukt (HTTP {resp.status_code})"
        soup = BeautifulSoup(resp.text, "lxml")
        for sc in soup.find_all("script", type="application/ld+json"):
            try:
                data = json.loads(sc.string or "")
                if isinstance(data, list): data = data[0]
                if data.get("@type") == "Product":
                    offers = data.get("offers", {})
                    if isinstance(offers, list): offers = offers[0]
                    price = offers.get("price") or offers.get("lowPrice")
                    if price:
                        return float(str(price).replace(",", ".")), "Succes (JSON-LD)"
            except: pass
        el = soup.find(itemprop="price")
        if el:
            try: return float(str(el.get("content") or el.get_text()).replace(",", ".")), "Succes (Schema.org)"
            except: pass
        # Aldi: {"priceValue":2.49,...}  (ook met escaped quotes)
        for sc in soup.find_all("script"):
            m = re.search(r'priceValue.{1,6}?([\d]+\.[\d]{1,2})', sc.string or "")
            if m:
                try: return float(m.group(1)), "Succes (JSON-LD)"
                except: pass
        return None, "Prijs niet automatisch gevonden"
    except Exception as e:
        return None, f"Fout: {str(e)[:60]}"

def push_to_github(token: str):
    try:
        headers = {
            "Authorization": f"token {token}",
            "Accept": "application/vnd.github.v3+json",
        }
        # Haal huidige SHA op
        r = requests.get(f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_FILE}", headers=headers)
        sha = r.json().get("sha", "")
        # Lees het bestand
        with open(EXCEL_PATH, "rb") as f:
            content = base64.b64encode(f.read()).decode()
        payload = {
            "message": f"Nieuw product toegevoegd via app - {date.today()}",
            "content": content,
            "sha": sha,
        }
        r2 = requests.put(f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_FILE}", json=payload, headers=headers)
        return r2.status_code in (200, 201)
    except:
        return False

st.set_page_config(
    page_title="QP Price Monitor",
    page_icon="📊",
    layout="wide",
)

st.markdown("""
<style>
/* Multiselect tags: neutrale grijze stijl */
[data-baseweb="tag"] {
    background-color: #e8e8e8 !important;
    border: 1px solid #ccc !important;
    border-radius: 4px !important;
}
[data-baseweb="tag"] span[title] {
    color: #333 !important;
    font-size: 0.82rem !important;
}
[data-baseweb="tag"] svg {
    fill: #666 !important;
}
[data-baseweb="tag"]:hover {
    background-color: #d8d8d8 !important;
}
</style>
""", unsafe_allow_html=True)

EXCEL_PATH = Path(__file__).parent / "data.xlsx"

@st.cache_data(ttl=600)  # Herlaadt elke 10 minuten automatisch
def load_data():
    df = pd.read_excel(EXCEL_PATH, sheet_name="Sheet1")
    df["datum"] = pd.to_datetime(df["datum"], errors="coerce")
    df = df.dropna(subset=["datum"])
    df["prijs"] = pd.to_numeric(df["prijs"], errors="coerce")
    df.loc[df["prijs"] == 0, "prijs"] = None
    return df

df = load_data()

SUCCESS_STATUSES = [
    "Succes (Schema.org)", "Succes (JSON-LD)", "Succes (Visual)",
    "Succes", "Succes (HTML price)", "Succes (Visueel HTML)",
]
df_ok = df[df["status"].isin(SUCCESS_STATUSES) & df["prijs"].notna()]

# ── Sidebar: Nieuw product toevoegen ──────────────────────────────────────────
with st.sidebar.expander("➕ Nieuw product toevoegen"):
    new_url = st.text_input("Plak URL hier", key="new_url")

    if new_url:
        retailer_auto = detect_retailer(new_url)
        with st.spinner("Prijs ophalen..."):
            prijs_auto, status_auto = fetch_price_requests(new_url)

        if prijs_auto:
            st.success(f"Prijs gevonden: **€{prijs_auto:.2f}**")
        else:
            st.warning(f"{status_auto} — vul prijs handmatig in")

        with st.form("nieuw_product_form"):
            product_naam = st.text_input("Productnaam *")
            col_a, col_b = st.columns(2)
            with col_a:
                artikel_nr = st.text_input("Artikelnummer")
            with col_b:
                leverancier = st.selectbox("Leverancier", ["Queens", "Concurrent"])
            retailer_keuze = st.selectbox("Retailer", list(RETAILER_MAP.values()), index=list(RETAILER_MAP.values()).index(retailer_auto) if retailer_auto in RETAILER_MAP.values() else 0)
            prijs_input = st.number_input("Prijs (€)", value=float(prijs_auto) if prijs_auto else 0.0, min_value=0.0, format="%.2f")
            toevoegen = st.form_submit_button("✅ Toevoegen")

        if toevoegen:
            if not product_naam:
                st.error("Vul een productnaam in.")
            else:
                wb = load_workbook(EXCEL_PATH)
                ws = wb["Sheet1"]
                ws.append([
                    leverancier,
                    float(artikel_nr) if artikel_nr else None,
                    new_url,
                    product_naam.upper(),
                    prijs_input if prijs_input > 0 else None,
                    retailer_keuze,
                    status_auto or "Handmatig",
                    date.today(),
                    None,
                ])
                wb.save(EXCEL_PATH)

                token = st.secrets.get("GITHUB_TOKEN", "")
                if token:
                    ok = push_to_github(token)
                    if ok:
                        st.success(f"✅ **{product_naam}** toegevoegd en online gezet!")
                    else:
                        st.warning("Opgeslagen, maar GitHub-sync mislukt.")
                else:
                    st.success(f"✅ **{product_naam}** lokaal toegevoegd.")
                st.cache_data.clear()
                st.rerun()

st.sidebar.divider()

# ── Sidebar: Product verwijderen ───────────────────────────────────────────────
with st.sidebar.expander("🗑️ Product verwijderen"):
    df_uniek = df.drop_duplicates(subset=["product_naam", "retailer"])[["product_naam", "retailer"]].sort_values("product_naam")
    producten_lijst = [f"{r['product_naam']} — {r['retailer']}" for _, r in df_uniek.iterrows()]

    te_verwijderen = st.selectbox("Selecteer product × retailer", ["— kies —"] + producten_lijst, key="del_select")

    if te_verwijderen != "— kies —":
        prod_del, ret_del = te_verwijderen.rsplit(" — ", 1)
        aantal = len(df[(df["product_naam"] == prod_del) & (df["retailer"] == ret_del)])
        st.caption(f"Dit verwijdert **{aantal} meetpunten** uit de dataset.")

        bevestig = st.checkbox("Ja, ik weet het zeker", key="del_confirm")
        if st.button("🗑️ Verwijderen", disabled=not bevestig):
            df_nieuw = df[(df["product_naam"] != prod_del) | (df["retailer"] != ret_del)]

            # Schrijf terug naar Excel (alleen Sheet1 aanpassen, Analyse bewaren)
            wb = load_workbook(EXCEL_PATH)
            ws = wb["Sheet1"]
            # Wis alle data-rijen en schrijf opnieuw
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.value = None
            for i, (_, row) in enumerate(df_nieuw.iterrows(), start=2):
                ws.cell(i, 1, row["Leverancier"])
                ws.cell(i, 2, row["Artikel_nummer"])
                ws.cell(i, 3, row["URL"])
                ws.cell(i, 4, row["product_naam"])
                ws.cell(i, 5, row["prijs"])
                ws.cell(i, 6, row["retailer"])
                ws.cell(i, 7, row["status"])
                ws.cell(i, 8, row["datum"].date() if pd.notna(row["datum"]) else None)
                ws.cell(i, 9, row["Concurrentie"])
            wb.save(EXCEL_PATH)

            token = st.secrets.get("GITHUB_TOKEN", "")
            if token:
                ok = push_to_github(token)
                if ok:
                    st.success(f"✅ **{prod_del} — {ret_del}** verwijderd en online bijgewerkt!")
                else:
                    st.warning("Verwijderd, maar GitHub-sync mislukt.")
            else:
                st.success(f"✅ **{prod_del} — {ret_del}** lokaal verwijderd.")
            st.cache_data.clear()
            st.rerun()

st.sidebar.divider()

# ── Sidebar filters ────────────────────────────────────────────────────────────
st.sidebar.title("Filters")

retailers = sorted(df["retailer"].dropna().unique().tolist())

# 1. Product (eerst — afhankelijk van leverancier/retailer, dus eerst tijdelijk alles)
producten_alle = sorted(df_ok["product_naam"].dropna().unique().tolist())
sel_product = st.sidebar.selectbox("Product", ["— selecteer —"] + producten_alle)

# 2. Leverancier
leveranciers = ["Alle"] + sorted(df["Leverancier"].dropna().unique().tolist())
sel_leverancier = st.sidebar.selectbox("Leverancier", leveranciers)

# 3. Periode
datums = sorted(df["datum"].dt.date.unique())
datum_min, datum_max = datums[0], datums[-1]
sel_datum = st.sidebar.date_input(
    "Periode",
    value=(datum_min, datum_max),
    min_value=datum_min,
    max_value=datum_max,
)

# 4. Retailers — als expander net als nieuw product / verwijderen
with st.sidebar.expander("🏪 Retailers filteren", expanded=False):
    btn_col1, btn_col2 = st.columns(2)
    with btn_col1:
        if st.button("✓ Alle", use_container_width=True, key="ret_alle"):
            st.session_state["sel_retailers"] = retailers
    with btn_col2:
        if st.button("✗ Geen", use_container_width=True, key="ret_geen"):
            st.session_state["sel_retailers"] = []

    sel_retailers = st.multiselect(
        label="",
        options=retailers,
        default=st.session_state.get("sel_retailers", retailers),
        key="sel_retailers",
        label_visibility="collapsed",
    )

# ── Filtering toepassen ────────────────────────────────────────────────────────
filtered = df_ok.copy()
# ── Header ─────────────────────────────────────────────────────────────────────
st.title("📊 QP Price Monitor")
st.caption(f"Databron: {EXCEL_PATH.name}  ·  {len(df):,} meetpunten  ·  {df['datum'].dt.date.nunique()} meetmomenten")

# ── KPI cards ──────────────────────────────────────────────────────────────────
latest_date = df_ok["datum"].max()

# Bepaal de drempel: de 'vorige meting' moet minstens 30% van de huidige meting-omvang hebben
# zodat handmatig toegevoegde items (1 rij) niet als 'vorig meetmoment' worden gezien
latest_count = (df_ok["datum"] == latest_date).sum()
min_threshold = max(5, int(latest_count * 0.3))

date_counts = df_ok.groupby("datum").size()
full_prev_dates = sorted([
    d for d in date_counts.index
    if d < latest_date and date_counts[d] >= min_threshold
])
prev_date = full_prev_dates[-1] if full_prev_dates else None

col1, col2, col3, col4 = st.columns(4)

with col1:
    n_products = df_ok["product_naam"].nunique()
    st.metric("Producten gemonitord", n_products)

with col2:
    n_retailers = df_ok["retailer"].nunique()
    st.metric("Supermarkten", n_retailers)

with col3:
    success_rate = df[df["status"].isin(SUCCESS_STATUSES)].shape[0] / len(df) * 100
    st.metric("Scrape-succesrate", f"{success_rate:.1f}%")

with col4:
    st.metric("Laatste meting", latest_date.strftime("%d-%m-%Y"))

st.divider()

# ── Tabs ───────────────────────────────────────────────────────────────────────
tab4, tab1, tab2, tab3, tab5, tab6 = st.tabs([
    "🔔 Prijswijzigingen", "📈 Prijsverloop", "🏷️ Marktoverzicht", "🏪 Retailervergelijking", "🔍 Datakwaliteit", "📋 Alle data"
])

# ── Tab 1: Prijsverloop ────────────────────────────────────────────────────────
with tab1:
    if sel_product == "— selecteer —":
        st.info("Selecteer een product in de sidebar om het prijsverloop te bekijken.")
    else:
        prod_df = filtered[filtered["product_naam"] == sel_product].copy()
        prod_df = prod_df.sort_values("datum")

        fig = px.line(
            prod_df,
            x="datum",
            y="prijs",
            color="retailer",
            markers=True,
            title=f"Prijsverloop — {sel_product}",
            labels={"datum": "Datum", "prijs": "Prijs (€)", "retailer": "Retailer"},
        )
        fig.update_layout(hovermode="x unified", yaxis_tickprefix="€")
        st.plotly_chart(fig, use_container_width=True)

        # Prijsmutaties tabel
        st.subheader("Prijsmutaties")
        if prev_date is not None:
            cur = prod_df[prod_df["datum"] == latest_date][["retailer", "prijs"]].rename(columns={"prijs": "Nu"})
            prv = prod_df[prod_df["datum"] == prev_date][["retailer", "prijs"]].rename(columns={"prijs": "Vorige"})
            mut = cur.merge(prv, on="retailer", how="outer").sort_values("retailer")
            mut["Verschil"] = mut["Nu"] - mut["Vorige"]
            mut["Δ%"] = (mut["Verschil"] / mut["Vorige"] * 100).round(1)
            mut["Nu"] = mut["Nu"].apply(lambda x: f"€{x:.2f}" if pd.notna(x) else "—")
            mut["Vorige"] = mut["Vorige"].apply(lambda x: f"€{x:.2f}" if pd.notna(x) else "—")
            mut["Verschil"] = mut["Verschil"].apply(lambda x: f"{x:+.2f}" if pd.notna(x) else "—")
            mut["Δ%"] = mut["Δ%"].apply(lambda x: f"{x:+.1f}%" if pd.notna(x) else "—")
            st.dataframe(mut.reset_index(drop=True), use_container_width=True, hide_index=True)
        else:
            st.write("Niet genoeg meetmomenten voor vergelijking.")

# ── Tab 2: Marktoverzicht per categorie ───────────────────────────────────────
with tab2:
    # Categorie-indeling op basis van trefwoorden in productnaam (lowercase)
    CATEGORIE_MAP = {
        "🐟 Pangasius & Pangalicious": ["pangasi", "pangalicious", "pangafilet", "golden seafood"],
        "🐟 Wilde Zalm": ["zalm", "zalmfilet"],
        "🐟 Koolvis": ["koolvis"],
        "🐟 Kabeljauw": ["kabeljauw"],
        "🐟 Tilapia": ["tilapia"],
        "🧊 IJs & Crushed Ice": [" ice ", "ice balls", "ice cubes", "crushed ice", "ijsblok", "spinner"],
        "🥖 Knoflookbaguette": ["baguette"],
        "🥟 Orien Bites": ["orien", "gyoza", "gua bao"],
    }

    def get_categorie(naam: str) -> str:
        naam_lower = f" {naam.lower()} "
        for cat, trefwoorden in CATEGORIE_MAP.items():
            if any(t in naam_lower for t in trefwoorden):
                return cat
        return "📦 Overig"

    # Gebruik per product × retailer de meest recente prijs (niet gefixeerd op één datum)
    # Zo blijven alle producten zichtbaar, ook als ze op verschillende datums zijn toegevoegd
    meest_recent = (
        df_ok.sort_values("datum")
        .groupby(["product_naam", "retailer"])
        .last()
        .reset_index()[["Leverancier", "product_naam", "retailer", "datum", "prijs"]]
    )

    # Vorige prijs: op voet van de op-één-na-laatste datum per product × retailer
    vorige_prijs = (
        df_ok[df_ok["datum"] < df_ok.groupby(["product_naam", "retailer"])["datum"].transform("max")]
        .sort_values("datum")
        .groupby(["product_naam", "retailer"])
        .last()
        .reset_index()[["product_naam", "retailer", "datum", "prijs"]]
        .rename(columns={"prijs": "Vorige prijs", "datum": "Vorige datum"})
    )

    markt = meest_recent.merge(vorige_prijs, on=["product_naam", "retailer"], how="left")
    markt = markt.rename(columns={"prijs": "Laatste prijs", "datum": "Laatste datum"})
    markt["Δ (€)"] = markt["Laatste prijs"] - markt["Vorige prijs"]

    prijs_nu  = "Laatste prijs"
    prijs_prv = "Vorige prijs"

    markt["Categorie"] = markt["product_naam"].apply(get_categorie)

    # Sorteer: Queens boven, daarna alfabetisch op product en retailer
    markt["_sort_lev"] = markt["Leverancier"].map({"Queens": 0, "Concurrent": 1}).fillna(2)
    markt = markt.sort_values(["Categorie", "_sort_lev", "product_naam", "retailer"])

    st.subheader("Marktoverzicht per productcategorie")
    st.caption(f"Meest recente prijs per product × retailer, vergeleken met de vorige meting")

    alle_cats = sorted(markt["Categorie"].unique().tolist())

    for cat in alle_cats:
        cat_df = markt[markt["Categorie"] == cat].copy()

        with st.expander(f"{cat}  ({len(cat_df)} items)", expanded=True):
            display = cat_df[["Leverancier", "product_naam", "retailer", "Laatste datum", "Vorige prijs", "Laatste prijs", "Δ (€)"]].copy()
            display = display.rename(columns={
                "product_naam": "Product",
                "retailer": "Retailer",
                "Laatste datum": "Datum",
                "Vorige prijs": "Vorige prijs",
                "Laatste prijs": "Laatste prijs",
            })

            # Opmaak
            display["Datum"]         = display["Datum"].dt.strftime("%d-%m-%Y")
            display["Vorige prijs"]  = display["Vorige prijs"].apply(lambda x: f"€{x:.2f}" if pd.notna(x) else "—")
            display["Laatste prijs"] = display["Laatste prijs"].apply(lambda x: f"€{x:.2f}" if pd.notna(x) else "—")
            display["Δ (€)"]         = display["Δ (€)"].apply(
                lambda x: (f"+€{x:.2f}" if x > 0 else f"-€{abs(x):.2f}") if pd.notna(x) and x != 0 else ("—" if pd.isna(x) else "")
            )

            st.dataframe(display.reset_index(drop=True), use_container_width=True, hide_index=True)

# ── Tab 3: Retailervergelijking ────────────────────────────────────────────────
with tab3:
    if sel_product == "— selecteer —":
        st.info("Selecteer een product in de sidebar voor de retailervergelijking.")
    else:
        prod_latest = df_ok[
            (df_ok["product_naam"] == sel_product) & (df_ok["datum"] == latest_date)
        ].copy()

        if prod_latest.empty:
            st.warning("Geen recente data beschikbaar voor dit product.")
        else:
            prod_latest = prod_latest.sort_values("prijs")
            fig3 = px.bar(
                prod_latest,
                x="retailer",
                y="prijs",
                color="Leverancier",
                title=f"Prijsvergelijking per retailer — {sel_product}",
                labels={"prijs": "Prijs (€)", "retailer": "Retailer"},
                color_discrete_map={"Queens": "#0068c9", "Concurrent": "#ff6b35"},
                text_auto=".2f",
            )
            fig3.update_traces(texttemplate="€%{y:.2f}")
            fig3.update_layout(yaxis_tickprefix="€")
            st.plotly_chart(fig3, use_container_width=True)

            # Goedkoopste / duurste
            c1, c2 = st.columns(2)
            with c1:
                cheapest = prod_latest.loc[prod_latest["prijs"].idxmin()]
                st.success(f"**Goedkoopst:** {cheapest['retailer']} — €{cheapest['prijs']:.2f}")
            with c2:
                most_exp = prod_latest.loc[prod_latest["prijs"].idxmax()]
                st.error(f"**Duurste:** {most_exp['retailer']} — €{most_exp['prijs']:.2f}")

# ── Tab 4: Prijswijzigingen ────────────────────────────────────────────────────
with tab4:
    st.subheader(f"Prijswijzigingen: {prev_date.strftime('%d-%m-%Y') if prev_date else '—'} → {latest_date.strftime('%d-%m-%Y')}")

    if prev_date is None:
        st.info("Niet genoeg meetmomenten voor vergelijking.")
    else:
        latest_df = df_ok[df_ok["datum"] == latest_date][["Leverancier", "product_naam", "retailer", "prijs"]].rename(columns={"prijs": "Nieuwe prijs"})
        prev_df   = df_ok[df_ok["datum"] == prev_date][["product_naam", "retailer", "prijs"]].rename(columns={"prijs": "Oude prijs"})

        wijzigingen = latest_df.merge(prev_df, on=["product_naam", "retailer"], how="left")
        wijzigingen["Verschil (€)"] = wijzigingen["Nieuwe prijs"] - wijzigingen["Oude prijs"]
        wijzigingen["Verschil (%)"] = (wijzigingen["Verschil (€)"] / wijzigingen["Oude prijs"] * 100).round(1)

        # Filter opties
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            filter_wijz = st.selectbox("Toon", ["Alle wijzigingen", "Prijsstijgingen", "Prijsdalingen"], key="wijz_filter")
        with col_f2:
            filter_lev = st.selectbox("Leverancier", ["Alle", "Queens", "Concurrent"], key="wijz_lev")

        # Standaard: alleen producten met een echte prijswijziging
        toon = wijzigingen[wijzigingen["Verschil (€)"].notna() & (wijzigingen["Verschil (€)"] != 0)].copy()
        if filter_lev != "Alle":
            toon = toon[toon["Leverancier"] == filter_lev]
        if filter_wijz == "Prijsstijgingen":
            toon = toon[toon["Verschil (€)"] > 0]
        elif filter_wijz == "Prijsdalingen":
            toon = toon[toon["Verschil (€)"] < 0]

        toon = toon.sort_values("Verschil (€)", ascending=False)

        # KPI's
        n_stijging = (wijzigingen["Verschil (€)"] > 0).sum()
        n_daling   = (wijzigingen["Verschil (€)"] < 0).sum()
        n_gelijk   = (wijzigingen["Verschil (€)"] == 0).sum()
        k1, k2, k3 = st.columns(3)
        k1.metric("📈 Gestegen", n_stijging)
        k2.metric("📉 Gedaald", n_daling)
        k3.metric("➡️ Ongewijzigd", n_gelijk)

        # Tabel
        col_oud   = f"Oude prijs ({prev_date.strftime('%d-%m-%Y')})"
        col_nieuw = f"Nieuwe prijs ({latest_date.strftime('%d-%m-%Y')})"

        toon_display = toon.copy()
        toon_display["Oude prijs"]   = toon_display["Oude prijs"].apply(lambda x: f"€{x:.2f}" if pd.notna(x) else "—")
        toon_display["Nieuwe prijs"] = toon_display["Nieuwe prijs"].apply(lambda x: f"€{x:.2f}" if pd.notna(x) else "—")
        toon_display["Verschil (€)"] = toon_display["Verschil (€)"].apply(lambda x: f"{x:+.2f}" if pd.notna(x) else "nieuw")
        toon_display["Verschil (%)"] = toon_display["Verschil (%)"].apply(lambda x: f"{x:+.1f}%" if pd.notna(x) else "—")
        toon_display = toon_display.rename(columns={
            "product_naam": "Product",
            "retailer": "Retailer",
            "Oude prijs": col_oud,
            "Nieuwe prijs": col_nieuw,
        })

        st.dataframe(
            toon_display[["Leverancier", "Product", "Retailer", col_oud, col_nieuw, "Verschil (€)", "Verschil (%)"]].reset_index(drop=True),
            use_container_width=True,
            hide_index=True,
        )

        # ── Excel download ────────────────────────────────────────────────────
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        datum_oud   = f"{prev_date.day}-{prev_date.month}-{prev_date.year}"
        datum_nieuw = f"{latest_date.day}-{latest_date.month}-{latest_date.year}"

        # Snapshot van de huidige data (voorkomt stale-buffer bug)
        export_rows = toon[["Leverancier", "product_naam", "retailer",
                             "Oude prijs", "Nieuwe prijs",
                             "Verschil (€)", "Verschil (%)"]].reset_index(drop=True).copy()

        wb = Workbook()
        ws = wb.active
        ws.title = "Prijswijzigingen"

        # Stijlen
        HDR_FILL  = PatternFill("solid", fgColor="1F4E79")   # donkerblauw
        HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
        HDR_ALIGN = Alignment(wrap_text=True, vertical="bottom", horizontal="center")
        thin      = Side(style="thin", color="CCCCCC")
        BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)
        FILL_ODD  = PatternFill("solid", fgColor="EBF3FB")   # lichtblauw
        FILL_EVEN = PatternFill("solid", fgColor="FFFFFF")

        euro_fmt = '_-"€"* #,##0.00_-;-"€"* #,##0.00_-;_-"€"* "-"??_-;_-@_-'
        pct_fmt  = '0.00%'

        # Header rij
        headers = [
            "Leverancier", "Product", "Retailer",
            f"Oude prijs\n{datum_oud}",
            f"Nieuwe prijs\n{datum_nieuw}",
            "Verschil (€)", "Verschil (%)",
        ]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font  = HDR_FONT
            cell.fill  = HDR_FILL
            cell.alignment = HDR_ALIGN
            cell.border = BORDER
        ws.row_dimensions[1].height = 32

        # Data rijen
        for r, row in enumerate(export_rows.itertuples(index=False), 2):
            fill = FILL_ODD if r % 2 == 0 else FILL_EVEN

            for c, val in enumerate([row[0], row[1], row[2]], 1):
                cell = ws.cell(r, c, val)
                cell.fill = fill
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center")

            for c, val in [(4, row[3]), (5, row[4]), (6, row[5])]:
                cell = ws.cell(r, c, float(val) if pd.notna(val) else None)
                cell.number_format = euro_fmt
                cell.fill = fill
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center")

            pct_cell = ws.cell(r, 7, float(row[6]) / 100 if pd.notna(row[6]) else None)
            pct_cell.number_format = pct_fmt
            pct_cell.fill = fill
            pct_cell.border = BORDER
            pct_cell.alignment = Alignment(vertical="center", horizontal="right")

        # Kolombreedtes, bevroren header, auto-filter
        for col, breedte in zip("ABCDEFG", [14, 46, 13, 18, 18, 14, 13]):
            ws.column_dimensions[col].width = breedte
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:G{len(export_rows) + 1}"

        # Bytes vastleggen (niet BytesIO meegeven — dat veroorzaakte de stale-bug)
        buf = io.BytesIO()
        wb.save(buf)
        excel_bytes = buf.getvalue()

        bestandsnaam = f"prijswijzigingen_{prev_date.strftime('%Y%m%d')}_{latest_date.strftime('%Y%m%d')}.xlsx"
        st.download_button(
            label="⬇️ Download als Excel",
            data=excel_bytes,
            file_name=bestandsnaam,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# ── Tab 5: Datakwaliteit ───────────────────────────────────────────────────────
with tab5:
    st.subheader("Scrape-status overzicht")

    status_counts = df.groupby(["datum", "status"]).size().reset_index(name="aantal")
    fig4 = px.bar(
        status_counts,
        x="datum",
        y="aantal",
        color="status",
        title="Scrape-resultaten per meetmoment",
        labels={"datum": "Datum", "aantal": "Aantal", "status": "Status"},
    )
    st.plotly_chart(fig4, use_container_width=True)

    st.subheader("Mislukte / ontbrekende metingen")
    failed = df[~df["status"].isin(SUCCESS_STATUSES)][
        ["datum", "Leverancier", "product_naam", "retailer", "status"]
    ].sort_values("datum", ascending=False)
    st.dataframe(failed.reset_index(drop=True), use_container_width=True, hide_index=True)

    st.subheader("Meest recente meting per product × retailer")
    coverage = (
        df_ok.groupby(["product_naam", "retailer"])["datum"]
        .max()
        .reset_index()
        .rename(columns={"datum": "Laatste meting"})
    )
    coverage["Laatste meting"] = coverage["Laatste meting"].dt.strftime("%d-%m-%Y")
    st.dataframe(coverage, use_container_width=True, hide_index=True)

# ── Tab 6: Alle data ───────────────────────────────────────────────────────────
with tab6:
    st.subheader("Alle meetdata")

    tbl_df = filtered.copy()

    col_f1, col_f2, col_f3 = st.columns(3)
    with col_f1:
        zoek = st.text_input("Zoek op productnaam", "")
    with col_f2:
        status_opties = ["Alle"] + sorted(tbl_df["status"].dropna().unique().tolist())
        sel_status = st.selectbox("Status", status_opties, key="tbl_status")
    with col_f3:
        sort_kolom = st.selectbox("Sorteren op", ["datum", "product_naam", "retailer", "prijs"], key="tbl_sort")

    if zoek:
        tbl_df = tbl_df[tbl_df["product_naam"].str.contains(zoek, case=False, na=False)]
    if sel_status != "Alle":
        tbl_df = tbl_df[tbl_df["status"] == sel_status]

    tbl_df = tbl_df.sort_values(sort_kolom, ascending=(sort_kolom != "datum"))
    tbl_display = tbl_df[["datum", "Leverancier", "Artikel_nummer", "product_naam", "prijs", "retailer", "status"]].copy()
    tbl_display["datum"] = tbl_display["datum"].dt.strftime("%d-%m-%Y")
    tbl_display["prijs"] = tbl_display["prijs"].apply(lambda x: f"€{x:.2f}" if pd.notna(x) else "—")

    st.caption(f"{len(tbl_display):,} rijen")
    st.dataframe(tbl_display.reset_index(drop=True), use_container_width=True, hide_index=True)

    csv = tbl_df.to_csv(index=False, sep=";", decimal=",").encode("utf-8-sig")
    st.download_button(
        label="⬇️ Download als CSV",
        data=csv,
        file_name="price_monitor_export.csv",
        mime="text/csv",
    )
